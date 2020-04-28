#在miao_collect中使用
# 整理评分/价格？类似readfile

from openpyxl import load_workbook


def main():
    for excel in range(1, 121):
        print('正在改第' + str(excel) + '个表格')
        readbook = load_workbook('E:\大创\价格数据\京东女装_价格\女装' + str(excel) + '.xlsx')
        rdsheet = readbook['Sheet1']
        writebook = load_workbook('E:\大创\汇总数据处理后\女装评分\女装_' + str(excel) + '.xlsx')
        wtsheet = writebook['Sheet']
        wtsheet.cell(1, 3, '比价网')
        wtsheet.cell(1, 4, '喵喵折')
        wtsheet.cell(1, 5, '营销策略')
        print(rdsheet.cell(4, 1).value)
        print(type(rdsheet.cell(4, 1).value))

        for val in range(2, wtsheet.max_row + 1):
            if '2017' in wtsheet.cell(val, 1).value:
                for i in range(4, rdsheet.max_row + 1):
                    if wtsheet.cell(val, 1).value == str(rdsheet.cell(i, 1).value)[:-9].replace('-', '/').replace('/0', '/'):
                        wtsheet.cell(val, 3, rdsheet.cell(i, 2).value)
                        wtsheet.cell(val, 4, rdsheet.cell(i, 3).value)
                        wtsheet.cell(val, 5, rdsheet.cell(i, 4).value)
            else:
                pass

        writebook.save('E:\大创\汇总数据处理后\女装评分\女装_' + str(excel) + '.xlsx')


if __name__ == '__main__':
    main()
