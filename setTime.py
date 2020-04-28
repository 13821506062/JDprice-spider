# 把2017-01-01改成2017/1/1时间格式

from datetime import date
from openpyxl import load_workbook


def main():
    for excel in range(1, 151):
        print('正在改第' + str(excel) + '个表格')
        readbook = load_workbook('E:\大创\汇总数据处理后\洗发水评分\洗发水_' + str(excel) + '.xlsx')
        sheet = readbook['Sheet']
        da = []
        for item in sheet['A']:
            if item.value == '':
                break
            if '-' not in item.value:
                pass
            else:

                ls = item.value.split('-')
                d = date(int(ls[0]), int(ls[1]), int(ls[2]))
                da.append(d)
        val = 3
        for item in da:
            sheet.cell(val, 1, item).number_format = 'YYYY/M/D'
            val = val + 1
        readbook.save('E:\大创\汇总数据处理后\洗发水评分\洗发水_' + str(excel) + '.xlsx')


if __name__ == '__main__':
    main()
